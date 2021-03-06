# 测定杨氏模量实验的数据处理脚本
# 图形界面待开发

import math
import openpyxl as opx

p = math.pi


def get_by_io():
    # 对标尺数据的处理
    # 获取最基本的xi+和xi-的数据
    a = input('请输入xi+的九个值，单位mm，每个值之间用空格隔开，不要有其他额外输入').split()
    b = input('请输入xi-的九个值，单位mm，每个值之间用空格隔开，不要有其他额外输入').split()

    # 对直径数据的处理
    # 获取最基本的d的数据
    c = input('请输入0.00kg时测得的三个钢丝直径的数据，单位mm，各数据间用空格隔开，不要有其他额外输入').split()
    d = input('请输入9.00kg时测得的三个钢丝直径的数据，单位mm，各数据间用空格隔开，不要有其他额外输入').split()
    return a, b, c, d


def get_by_xlsx():
    workbook = opx.load_workbook('./1.xlsx')
    sheet = workbook.active
    a = []
    b = []
    c = []
    d = []
    for i in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=10):
        for k in i:
            a.append(k.value)
    for i in sheet.iter_rows(min_row=4, max_row=4, min_col=1, max_col=10):
        for k in i:
            b.append(k.value)
    for i in sheet.iter_rows(min_row=6, max_row=6, min_col=1, max_col=3):
        for k in i:
            c.append(k.value)
    for i in sheet.iter_rows(min_row=8, max_row=8, min_col=1, max_col=3):
        for k in i:
            d.append(k.value)
    print(a, b, c, d)
    return a, b, c, d


def main(xiPlus, xiMinus, d_0, d_9):
    # 以下是我自己测得的实验数据
    # xiPlus = [6.1, 9.2, 12.1, 15.2, 18.6, 21.4, 24.3, 27.8, 30.3, 33.8]
    # xiMinus = [5.2, 8.7, 12.0, 14.9, 17.8, 20.9, 24.0, 27.2, 30.7, 33.0]
    # d_0 = [0.630, 0.632, 0.634]
    # d_9 = [0.630, 0.628, 0.627]
    # L = 738
    # H = 812
    # b = 501

    xi = []
    for i in range(len(xiPlus)):
        xi.append((float(xiPlus[i]) + float(xiMinus[i])) / 2)

    # 进行初步数据处理
    delta_xi = []
    for i in range(5):
        delta_xi.append(xi[5 + i] - xi[i])

    # 计算各个均值
    average_delta_xi = 0
    for x in delta_xi:
        average_delta_xi = average_delta_xi + x
    average_delta_xi = average_delta_xi / 5
    AD_xi = 0
    for y in delta_xi:
        AD_xi = AD_xi + abs(y - average_delta_xi)
    AD_xi = AD_xi / 5

    # 对数据进行进一步处理
    average_d = []  # 对上中下三个尺寸取均值
    for i in range(len(d_0)):
        average_d.append((float(d_0[i]) + float(d_9[i])) / 2)
    d_average = 0
    for x in average_d:
        d_average = d_average + x
    d_average = d_average / 3
    delta_d = []
    average_delta_d = 0
    for x in average_d:
        delta_d.append(abs(x - d_average))
        average_delta_d = average_delta_d + x
    average_delta_d = average_delta_d / 3

    b = float(input('请输入b的值，单位mm'))
    L = float(input('请输入L的值，单位mm'))
    H = float(input('请输入H的值，单位mm'))
    F = 5 * 9.80

    Y = (8.0 * F * L * H) / (p * (d_average ** 2) * b * average_delta_d)

    for i in range(len(xi)):
        xi[i] = round(float(xi[i]), 3)
    for i in range(len(delta_xi)):
        delta_xi[i] = round(delta_xi[i], 3)
    for i in range(len(d_0)):
        average_d[i] = round(average_d[i], 3)
        delta_d[i] = round(delta_d[i], 3)

    print('平均标尺刻度为', xi)
    print('标尺刻度改变量为', delta_xi)
    print('标尺该变量平均值为{:.3f}'.format(average_delta_xi), 'mm')
    print('标尺刻度该变量的平均偏差为{:.3f}'.format(AD_xi), 'mm')
    print('钢丝直径平均值分别为', average_d)
    print('钢丝直径平均值为{:.3f}'.format(d_average), 'mm')
    print('钢丝直径该变量为', delta_d, )
    print('钢丝直径改变量的平均偏差为{:.3f}'.format(average_delta_d), 'mm')
    print('杨氏模量的值为{:.3f}'.format(Y), 'N/mm2')


if __name__ == '__main__':
    print('=======================================================')
    print('|    测定金属杨氏模量实验的数据处理脚本，一站式解决所有计算   |')
    print('|                      made by zzx                    |')
    print('---------------------------------------version 1.2-----')
    print('|   you can also find the code at the following URL:  |')
    print('| https://github.com/coder-Zzx/experiment_project.git |')
    print('=======================================================', end='\n\n')
    print('=======================================================')
    print('|                    1.从键盘输入                      |')
    print('|                    2.从Excel文件输入                 |')
    ans = input('|\t请选择输入模式\t|')
    if ans == '1':
        xiPlus, xiMinus, d_0, d_9 = get_by_io()
    if ans == '2':
        print('请在代码相同路径下的1.xlsx工作表内按格式输入实验数据.')
        xiPlus, xiMinus, d_0, d_9 = get_by_xlsx()
    main(xiPlus=xiPlus, xiMinus=xiMinus, d_0=d_0, d_9=d_9)
