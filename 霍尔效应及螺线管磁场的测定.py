from matplotlib import pyplot as plt
import numpy as np
import openpyxl as opx
import math

p = math.pi
u = 4 * p * (10 ** (-7))
N = 3019
L = 0.276
D = ((40 + 56) / 2) * (10 ** (-3))
B = (u * N * 0.6) / ((L ** 2 + D ** 2) ** 0.5)
e = 1.6 * (10 ** (-19))
d = 0.3 * (10 ** (-3))


def func():
    wb = opx.load_workbook("1.xlsx", data_only=True)
    ws = wb["Sheet" + str(1)]
    fig, axis = plt.subplots(ncols=1, nrows=3)
    plt.subplots_adjust(left=0.1, wspace=0.5, hspace=0.5)
    Us = []
    ICH = []
    _row = 3
    _column = 1
    _column_ = 6
    while True:
        if _row > 12:
            break
        Us.append(ws.cell(row=_row, column=6).value)
        ICH.append(ws.cell(row=_row, column=1).value)
        _row += 1
    Us_np = np.array(Us)
    ICH_np = np.array(ICH)
    axis[0].plot(ICH_np, Us_np, color='red', linewidth=1, linestyle='-', label=r'$U_H-I_{CH}\ relation\ curve$')
    for i in range(len(Us)):
        axis[0].scatter([ICH[i]], Us[i], s=30)
    axis[0].set_xlabel('$I_{CH}/mA$')
    axis[0].set_ylabel('$U_H/mV$')
    axis[0].legend()
    slope, i = np.polyfit(Us_np, ICH_np, 1)
    KH = slope / B
    print('KH=', KH)
    n = 1 / (KH * e * d)
    print('n=', n)
    IM_np = ICH_np
    UH1 = []
    _row = 16
    while True:
        if _row > 25:
            break
        UH1.append(ws.cell(row=_row, column=6).value)
        _row += 1
    UH1_np = np.array(UH1)
    axis[1].plot(IM_np, UH1_np, color='blue', linewidth=1, linestyle='--', label=r'$U_H-I_M\ relation\ curve$')
    for i in range(len(Us)):
        axis[1].scatter([ICH[i]], UH1[i], s=30)
    axis[1].set_xlabel('$I_M/mA$')
    axis[1].set_ylabel('$U_H/mV$')
    axis[1].legend()

    _r = 40
    _column1 = 10
    UH2 = []
    B1 = []
    X = []
    while True:
        UH0 = ws.cell(row=_r, column=10).value
        UH2.append(UH0)
        Ich = 5 ** (-3)
        B0 = UH0 / (KH * Ich)
        ws.cell(row=_r, column=12, value=B0)
        B1.append(B0)
        _r -= 1
        if _r < 30:
            break
    while True:
        _r += 1
        UH0 = ws.cell(row=_r, column=11).value
        UH2.append(UH0)
        Ich = 5 ** (-3)
        B0 = UH0 / (KH * Ich)
        ws.cell(row=_r, column=13, value=B0)
        B1.append(B0)
        if _r >= 40:
            break
    while True:
        X0 = -ws.cell(row=_r, column=1).value
        X.append(X0)
        _r -= 1
        if _r < 30:
            break
    while True:
        _r += 1
        X0 = ws.cell(row=_r, column=1).value
        X.append(X0)
        if _r >= 40:
            break
    UH2_np = np.array(UH2)
    X_np = np.array(X)
    B_np = np.array(B1)
    axis[2].plot(X_np, B_np, color='green', linewidth=1, linestyle='-.', label=r'$B-X\ relation\ curve$')
    axis[2].set_xlabel('$X/mx$')
    axis[2].set_ylabel('$B/mT$')
    for i in range(len(X)):
        axis[2].scatter([X[i]], B1[i], s=30)

    print(B1)
    print(X)

    wb.save("1.xlsx")

    plt.legend()
    plt.show()


func()
